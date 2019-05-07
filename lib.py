import numpy as np
import openpyxl as xl
from pickle import load
from math import inf
from SequenceParser import SequenceParser

######################################
##		Miscellaneous Functions		##
######################################

def distance_with_len(seq1, seq2):
	# print(f"Comparing {seq1} and {seq2}")
	if len(seq1) >= len(seq2):
		longerSeq = seq1
		shorterSeq = seq2
	else:
		longerSeq = seq2
		shorterSeq = seq1
	prefix = 0
	for i, n in enumerate(shorterSeq):
		if shorterSeq[i] == longerSeq[i]:
			prefix += 1
	if not prefix:
		return 1
	suffix = len(longerSeq) - prefix
	distance = (suffix/prefix)/len(longerSeq)
	return distance

def loadFromPickle(file,lenCutoff,countCutoff):
	with open(file,'rb') as file:
		insDict = load(file)
	wellDict = {}
	for ins in insDict:
		if not ins == 'ROOT':
			if len(ins) in range(lenCutoff[0],lenCutoff[1]):
				for well in insDict[ins]["counts"]:
					if insDict[ins]["counts"][well] >= countCutoff:
						if well not in wellDict:
							wellDict[well] = {}
							wellDict[well][ins] = {"insCount":0}
						else:
							wellDict[well][ins] = {"insCount":0}
						wellDict[well][ins]["insCount"] = insDict[ins]["counts"][well]
	return wellDict
	# group_1 = ["0","1","2","3","5","6","10","14"]
	# group_2 = ["4","7","8","9","11","12","13","15"]
	# for _ in group_2:
	# 	wellDict[f"well_{_}"] = {}
	# for ins in insDict:
	# 	if not ins == 'ROOT':
	# 		if len(ins) in range(lenCutoff[0],lenCutoff[1]):
	# 			for well in insDict[ins]["counts"]:
	# 				if well in wellDict:
	# 					if insDict[ins]["counts"][well] >= countCutoff:
	# 						wellDict[well][ins] = {"insCount":insDict[ins]["counts"][well]}
	# return wellDict


def loadFromWorkbook(file):
	wb = xl.load_workbook(filename=file)
	ws = wb['Sheet1']

	wellDict = {}
	for c in ws.columns:
		well = f'well_{c[0].value}'
		for i in range(ws.max_row):
			if i == 0:
				wellDict[well] = {}
			else:
				wellDict[well][str(c[i].value)] = {"insCount":51}
	return wellDict


def computeWellDistance(data):
	#need to unpack the data 
	cellsX = list(data["wellX"]) # All the cells in wellX (need to convert to list for reference issues)
	cellsY = list(data["wellY"]) # All the cells in wellY
	nameX = data["nameX"]
	nameY = data["nameY"]
	# Find the average prefix length between these two wells:
	print(f"Working on {nameX}-{nameY}",end='\r')
	allPrefixes = 0
	ctr = 0
	newWell = []
	for cellX in cellsX:
		maxPrefix = 0
		longestMatch = ""
		for cellY in cellsY:
			longerSeq, shorterSeq = (cellX, cellY) if len(cellX) >= len(cellY) else (cellY, cellX)
			prefix = 0
			newCell = ""
			for i in range(len(shorterSeq)):
				if not longerSeq[i] == shorterSeq[i]:
					break
				prefix += 1
				newCell = longerSeq[:i+1]
			if prefix > maxPrefix:
				maxPrefix = prefix
				longestMatch = newCell
		allPrefixes += maxPrefix
		ctr += 1
		newWell.append(longestMatch)
	# return allPrefixes/ctr, newWell #OLD
	return {"distance":allPrefixes/ctr,"nameX":nameX,"nameY":nameY}

def findClosestMatch(cellX,cellsY,ignore = [],memDict = {}):
	winning_seq = ""
	winning_distance = inf
	s = SequenceParser()
	#Iterate over all the sequences in well y and find the single highest match
	#But ignore cells that have already been checked.
	for cellY in [y for y in cellsY if y not in ignore]:
		if cellX+"-"+cellY in memDict:
			dist = memDict[cellX+"-"+cellY]
		else:
			dist = s.stringDist(cellX,cellY)
			memDict[cellX+"-"+cellY] = dist
		# print(f"Checking {cellX} against {cellY} with a distance of {dist}")
		if dist >= 0: #Prevent the -1s
			if dist < winning_distance:
				winning_seq = cellY
				winning_distance = dist
	if len(winning_seq) == 0:
		raise ValueError("Ah shit.")
		# return "",-1
	del s #free up memory
	return winning_seq, winning_distance

'''
Assign X Cell
	This function takes in a single X cell and returns one of three results:
		1.) If it finds a space in the tracker that is open, return None, positive distance
		2.) If it finds space, but had to kick out a cell to fit, return that oldCell and a negative distance
		3.) If it can't find space because all the Y's are taken, return the same cell and None distance.
'''
def assignXCell(cellX,cellsY,ignoreList,tracker,memDict,level=0):
	distance = 0
	while len(ignoreList) < len(cellsY): #inner loop is comparing a single cell in X to all cells in Y
		#findClosestMatch returns the closest cellY to cellX. But there might already be one closer.
		cellY,dist = findClosestMatch(cellX,cellsY,ignoreList, memDict)
		# so we check against the tracker dict, and if there is a closer one, we find the second closest Y, etc...
		if cellY in tracker:
			if len(tracker[cellY]) < level+1: #There isn't another level, so we can add this one
				tracker[cellY].append((cellX,dist))
				return None,dist #all done!
			if tracker[cellY][level][1] > dist: #Our current cell is closer
				oldCell = tracker[cellY][level][0]
				oldDist = tracker[cellY][level][1]
				distance -= oldDist - dist #take the difference of the distances, and subtract the total distance.
				# print(f"Putting {oldCell} back in the queue, replacing with {cellX}, updating total distance")
				tracker[cellY][level] = (cellX,dist)
				return oldCell,distance # (CASE 2) Return the kicked out cell, and the new negative distance
		else: #this is the first time we're seeing this cellY
			tracker[cellY] = [(cellX,dist)] #Don't worry about level here
			return None, dist # (CASE 1) No cell kicked out, this cell fit in the tracker. Positive distance
		ignoreList.append(cellY) #run the loop again, but ignore that cellY this time
	else:
		return cellX,None #(CASE 3) Reject it and return the original cell back
	raise RuntimeError("This shouldn't happen...")